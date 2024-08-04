import pandas as pd
import re
from time import time
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl import Workbook

# 文件路径（更具描述性的命名）
professional_details_path = "C:\\Users\\simon\\Desktop\\python-projects\\python-tools\\高考全国院校投档线-文件转换_数据清洗\\江西\\2024\\院校招生专业组专业明细.xlsx"
enrollment_statistics_path = "C:\\Users\\simon\\Desktop\\python-projects\\python-tools\\高考全国院校投档线-文件转换_数据清洗\\江西\\2024\\江西省2024年普通高校招生本科投档情况统计表(历史类、物理类、三校生类).xlsx"

# 计时开始
start_time = time()

# 读取文件
df1 = pd.read_excel(professional_details_path)
df2 = pd.read_excel(enrollment_statistics_path)

# 清洗数据，确保列名和数据一致
df1 = df1[["院校名称", "专业组", "选科", "包含专业"]]


def replace_brackets(text):
    # 使用正则表达式替换中文括号为英文括号
    return text.replace("（", "(").replace("）", ")")


# 将df1 中院校名称 的中文括号转换成英文
df1["院校名称"] = df1["院校名称"].apply(replace_brackets)
df2["院校名称"] = df2["院校名称"].apply(replace_brackets)

# 清洗df2中的专业组，移除“第”和“组”以及括号中的内容
df1["专业组"] = df1["专业组"].apply(lambda x: re.sub(r"\D", "", x).strip())

# 处理 NaN 值
df2["专业组名称"] = df2["专业组名称"].astype(str)
df2["专业组代码"] = df2["专业组名称"].apply(lambda x: re.sub(r"\D", "", x).strip())

# 创建唯一键
df1["key"] = df1["院校名称"] + "-" + df1["专业组"].astype(str)
df2["key"] = df2["院校名称"] + "-" + df2["专业组代码"].astype(str)

# 合并数据
df2 = df2.merge(
    df1[["key", "选科", "包含专业"]], on="key", how="left", suffixes=("", "_new")
)

# 用表1的值填充表2的对应列
df2["选科"] = df2["选科_new"].combine_first(df2["选科"])

df2["包含专业"] = df2["包含专业_new"].combine_first(df2["包含专业"])

# 删除临时列
df2.drop(columns=["选科_new", "包含专业_new", "key", "专业组代码"], inplace=True)

# 对位次列升序排序（处理可能的数据类型问题）
try:
    df2["位次"] = pd.to_numeric(df2["位次"], errors="coerce")
    df2.sort_values(by="位次", inplace=True)
except ValueError as e:
    print(f"Error sorting by '位次' column: {e}")


#  不要删除，不要识别，只是备注 保存 DataFrame 为 Excel 文件
# df2.to_excel(enrollment_statistics_path, index=False)

# # 使用 openpyxl 读取刚保存的 Excel 文件
# wb = load_workbook(enrollment_statistics_path)
# ws = wb.active

# # 设置边框样式（这里定义 thin_border）
# thin_border = Border(
#     left=Side(style="thin"),
#     right=Side(style="thin"),
#     top=Side(style="thin"),
#     bottom=Side(style="thin"),
# )

# 以下是对 Excel 文件进行格式设置的代码
wb = Workbook()
ws = wb.active
ws.title = "投档线"

# 将表头写入 Excel 工作表
header_row = df2.columns.values
for c in range(1, len(header_row) + 1):
    ws.cell(row=1, column=c).value = header_row[c - 1]

# 将数据写入 Excel 工作表（从第二行开始，避免覆盖表头）
for r in range(2, len(df2) + 2):
    for c in range(1, len(df2.columns) + 1):
        cell_value = df2.iloc[r - 2, c - 1]
        ws.cell(row=r, column=c).value = cell_value


# 获取包含专业列的列索引
col_index = df2.columns.get_loc("包含专业") + 1
# 包含专业列字母
special_col_letter = ws.cell(row=1, column=col_index).column_letter

# 合并循环进行格式设置
for col_idx, col in enumerate(ws.columns, start=1):
    column_letter = col[0].column_letter
    for cell in col:
        if column_letter == special_col_letter:
            cell.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    if column_letter == special_col_letter:
        ws.column_dimensions[column_letter].width = 70
    else:
        ws.column_dimensions[column_letter].width = 6

# 开启筛选（对所有列）
ws.auto_filter.ref = ws.dimensions

# 只是为了备注不需要处理，也不要删除，要保留这个注释 - 手动设置突出单元格规则  =AND(LEN(F1)<100,ISNUMBER(SEARCH("临床医学",F1)))

# 保存格式设置后的 Excel 文件
wb.save(enrollment_statistics_path)

# 计时结束
end_time = time()
print(f"处理时间: {end_time - start_time:.2f} 秒")
