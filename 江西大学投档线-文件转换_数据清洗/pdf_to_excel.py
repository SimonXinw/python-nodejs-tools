import os
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.dimensions import ColumnDimension
import time
from concurrent.futures import ThreadPoolExecutor

# 全局开始时间
global_start_time = time.time()

def read_pdf_with_pdfplumber(pdf_path):
    data = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if table:
                data.extend(table)
    return data

def adjust_column_width_and_alignment(ws, column_name, width, alignment):
    column_letter = None
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value and column_name in col[0].value:
            column_letter = col[0].column_letter
            break

    if column_letter:
        ws.column_dimensions[column_letter].width = width
        # 设置列头和列值对齐方式
        for col in ws[column_letter]:
            col.alignment = Alignment(
                horizontal=alignment if col.row != 1 else "center"
            )

def process_group(group_df, category, pdf_file, output_folder):
    # 按照投档线降序排序
    sorted_df = group_df.sort_values(by="投档线", ascending=False)

    # 删除已有的“序号”列（如果存在）
    sorted_df = sorted_df.drop(columns=["序号"], errors="ignore")

    # 添加新的序号列
    sorted_df.insert(0, "序号", range(1, len(sorted_df) + 1))

    # 确定输出文件名，修正文件名格式
    base_name = os.path.splitext(pdf_file)[0]
    output_file = os.path.join(output_folder, f"{base_name}_{category}.xlsx")

    # 写入Excel文件，禁止写入索引列
    sorted_df.to_excel(output_file, index=False, engine="openpyxl")

    # 调整“最低投档排名”列的宽度和对齐方式
    wb = load_workbook(output_file)
    ws = wb.active
    adjust_column_width_and_alignment(ws, "最低投档排名", 13, "center")
    adjust_column_width_and_alignment(ws, "院校名称", 32, "left")
    wb.save(output_file)

    # 记录结束时间并计算运行时间差
    end_time = time.time()
    elapsed_time = end_time - global_start_time
    print(f"文件 {pdf_file} 分类 {category} 处理消耗时间: {elapsed_time:.2f} 秒")

def process_pdf_data(pdf_data, pdf_file, pending_folder, output_folder):
    # 创建 DataFrame，不修改列名
    df = pd.DataFrame(pdf_data)

    # 确定中间输出文件名
    intermediate_file = os.path.join(pending_folder, f"{pdf_file.replace('.pdf', '.xlsx')}")

    # 写入中间Excel文件，禁止写入索引列
    df.to_excel(intermediate_file, index=False, header=False, engine="openpyxl")

    # 读取中间Excel文件，包括空数据
    df = pd.read_excel(intermediate_file, header=None)

    # 设置列名为第一行
    df.columns = df.iloc[0]
    df = df.drop(0)  # 删除第一行

    # 替换列名中的换行符
    df.columns = [col.replace("\n", "") for col in df.columns]

    # 确保 '投档线' 列存在
    if "投档线" in df.columns:
        # 将投档线列转换为数值类型
        df["投档线"] = pd.to_numeric(df["投档线"], errors="coerce")

        # 删除不完整的行，只保留有投档线的行
        df = df.dropna(subset=["投档线"])

        # 保存数据清洗后的中间结果
        df.to_excel(intermediate_file, index=False)

    else:
        print(f"在中间文件 '{intermediate_file}' 中未找到 '投档线' 列，跳过处理。")

    # 记录结束时间并计算运行时间差
    end_time = time.time()
    elapsed_time = end_time - global_start_time
    print(f"文件 {pdf_file} 数据清洗消耗时间: {elapsed_time:.2f} 秒")

def pdf_to_excel_and_process(pdf_folder, pending_folder, output_folder):
    # 创建输出文件夹
    if not os.path.exists(pending_folder):
        os.makedirs(pending_folder)
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # 获取 PDF 文件夹中的所有文件
    pdf_files = [f for f in os.listdir(pdf_folder) if f.endswith(".pdf")]

    with ThreadPoolExecutor(max_workers=4) as executor:
        for pdf_file in pdf_files:
            pdf_path = os.path.join(pdf_folder, pdf_file)
            pdf_data = read_pdf_with_pdfplumber(pdf_path)
            executor.submit(process_pdf_data, pdf_data, pdf_file, pending_folder, output_folder)

    print("PDF 文件转换和处理为 Excel 文件完成。")

def load_jiangxi_data(jiangxi_folder):
    jiangxi_data = {}
    for root, dirs, files in os.walk(jiangxi_folder):
        for file in files:
            if file.endswith(".xlsx"):
                path = os.path.join(root, file)
                df = pd.read_excel(path)
                if "院校名称" in df.columns:
                    for _, row in df.iterrows():
                        school_name = row["院校名称"]
                        jiangxi_data[school_name] = row.to_dict()
    return jiangxi_data

def append_school_data_to_excel(input_file, jiangxi_data):
    # 读取 Excel 文件
    df = pd.read_excel(input_file)

    # 将江西数据的列名添加到原来的 DataFrame 中
    additional_columns = list(next(iter(jiangxi_data.values())).keys())
    for col in additional_columns:
        if col not in df.columns:
            df[col] = None

    # 遍历“院校名称”列进行匹配数据
    if "院校名称" in df.columns:
        for i, row in df.iterrows():
            school_name = row["院校名称"]
            if school_name in jiangxi_data:
                matching_data = jiangxi_data[school_name]
                for key, value in matching_data.items():
                    if key in df.columns:
                        df.at[i, key] = value
    
    # 保存到同一个文件
    df.to_excel(input_file, index=False)

    return df  # 返回匹配后的数据框

def create_filtered_copy(input_file, output_folder):
    start_time = time.time()  # 记录开始时间

    # 读取 Excel 文件
    df = pd.read_excel(input_file)

    # 保存副本
    base_name = os.path.splitext(os.path.basename(input_file))[0]
    output_file = os.path.join(output_folder, f"默认筛选_{base_name}.xlsx")
    df.to_excel(output_file, index=False, engine="openpyxl")

    # 重新打开副本进行操作
    wb = load_workbook(output_file)
    ws = wb.active

    # 冻结首行
    ws.freeze_panes = ws['A2']

    # 添加筛选
    ws.auto_filter.ref = ws.dimensions

    # 默认筛选
    for cell in ws[1]:
        if cell.value == "科类":
            col_idx = cell.column
            ws.auto_filter.add_filter_column(col_idx - 1, ['理工'])
        if cell.value == "投档线":
            col_idx = cell.column
            ws.auto_filter.add_sort_condition(f"{cell.column_letter}2:{cell.column_letter}{ws.max_row}")

    # 保存文件
    wb.save(output_file)

    # 记录结束时间并计算运行时间差
    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f"文件 {input_file} 创建默认筛选副本并保存到 {output_file} 完成，消耗时间: {elapsed_time:.2f} 秒")

def process_pending_excels(pending_folder, output_folder, jiangxi_data):
    # 处理 pending_excel_folder 中的所有文件
    pending_files = [f for f in os.listdir(pending_folder) if f.endswith(".xlsx")]

    with ThreadPoolExecutor(max_workers=4) as executor:
        for pending_file in pending_files:
            input_path = os.path.join(pending_folder, pending_file)
            df = append_school_data_to_excel(input_path, jiangxi_data)
            create_filtered_copy(input_path, output_folder)

    # 再次处理 pending_excel_folder 中的文件
    for pending_file in pending_files:
        intermediate_file = os.path.join(pending_folder, pending_file)
        df = pd.read_excel(intermediate_file)
        
        # 确保 '投档线' 列存在
        if "投档线" in df.columns:
            # 按照科类分组并排序
            grouped = df.groupby("科类", sort=False)  # 不排序，保持原来的顺序

            for category, group_df in grouped:
                process_group(group_df, category, pending_file, output_folder)

if __name__ == "__main__":
    start_time = time.time()  # 记录开始时间

    pdf_folder = "pending_pdf"  # 替换成你的PDF文件夹路径
    pending_excel_folder = "pending_excel"  # 存放转换后中间Excel文件的文件夹路径
    output_excel_folder = "output_excel"  # 输出最终处理结果的文件夹路径

    # 加载江西文件夹中的数据
    jiangxi_folder = "江西"  # 替换成你的江西文件夹路径
    jiangxi_data = load_jiangxi_data(jiangxi_folder)

    # 将PDF转换为Excel文件并进行处理
    pdf_to_excel_and_process(pdf_folder, pending_excel_folder, output_excel_folder)

    # 匹配和追加学校数据，并创建默认筛选副本
    process_pending_excels(pending_excel_folder, output_excel_folder, jiangxi_data)

    end_time = time.time()  # 记录结束时间

    # 打印总耗时
    total_time = end_time - start_time
    print(f"总耗时: {total_time:.2f} 秒")
